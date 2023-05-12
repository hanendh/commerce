import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    rows = sheet.max_row
    return rows


def getColumnCount(file, sheetName):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    columns = sheet.max_row
    return columns


def ReadData(file, sheetName, row, column):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    return sheet.cell(row, column).value # return string data


def WriteData(file, sheetName, row, column, data):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    sheet.cell(row, column).value = data
    bk.save(file)


def GetGreenColor(file, sheetName, row, column):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    greenColor = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type='solid')
    sheet.cell(row, column).fill = greenColor
    bk.save(file)


def GetRedColor(file, sheetName, row, column):
    bk = openpyxl.load_workbook(file)
    sheet = bk[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(row, column).fill = redFill
    bk.save(file)
